import re
from openpyxl import load_workbook


excel_file = r"C:\Users\15322\Desktop\论文\论文数据\260410_京东_家用电器_大家电_空调_品类数据_月.xlsx"


def clean_text(value):
    if value is None:
        return ""
    text = str(value)
    text = text.replace("（", "(").replace("）", ")")
    text = text.replace("大一匹", "大1匹")
    text = text.replace("正一匹", "正1匹")
    text = text.replace("小一匹", "小1匹")
    text = text.replace("一匹", "1匹")
    text = text.replace("二匹", "2匹")
    text = text.replace("三匹", "3匹")
    text = text.replace("四匹", "4匹")
    text = text.replace("五匹", "5匹")
    text = re.sub(r"\s+", "", text)
    return text


def extract_hp(title):
    text = clean_text(title)

    hp_pattern = re.compile(
        r"(小|正|大)?(1\.5|1|2|3|5)\s*(匹|P|p)"
    )

    matches = list(hp_pattern.finditer(text))

    if matches:
        m = matches[-1]
        prefix = m.group(1)
        num = m.group(2)

        if num == "5":
            return "5 匹"

        if prefix == "小":
            return f"小 {num} 匹"
        elif prefix == "大":
            return f"大 {num} 匹"
        else:
            return f"正 {num} 匹"

    code_match = re.search(
        r"(?<!\d)(?:KFRD?|KFR|KF)?-?(20|22|23|25|26|33|35|46|48|50|51|60|72|75|120)(?:GW|LW|T2W)",
        text,
        re.I
    )

    if code_match:
        code = code_match.group(1)

        code_map = {
            "20": "小 1 匹",
            "22": "小 1 匹",
            "23": "小 1 匹",
            "25": "正 1 匹",
            "26": "大 1 匹",
            "33": "小 1.5 匹",
            "35": "正 1.5 匹",
            "46": "小 2 匹",
            "48": "正 2 匹",
            "50": "正 2 匹",
            "51": "大 2 匹",
            "60": "小 3 匹",
            "72": "正 3 匹",
            "75": "正 3 匹",
            "120": "5 匹"
        }

        return code_map.get(code, "")

    return ""


def extract_machine_type(title):
    text = clean_text(title)

    code_matches = re.findall(
        r"(?:KFRD?|KFR|KF)?-?\d{2,3}(GW|LW|T2W)",
        text,
        re.I
    )

    if code_matches:
        code = code_matches[-1].upper()

        if code == "GW":
            return 0
        elif code == "LW":
            return 1
        elif code == "T2W":
            return ""

    if re.search(r"柜机|立式|落地|立柜|柜式|圆柱", text):
        return 1

    if re.search(r"挂机|挂式|壁挂|壁挂式|挂壁|大挂机", text):
        return 0

    return ""


def extract_cold_warm_type(title):
    text = clean_text(title)

    if re.search(r"单冷|只冷|单制冷", text):
        return 0

    if re.search(r"(?<!R)KF-\d", text, re.I):
        return 0

    if re.search(r"KFR|冷暖|冷暖型|制冷暖|冷暖两用|制冷热|速冷暖|快速冷暖", text, re.I):
        return 1

    return ""


def extract_fresh_air(title):
    text = clean_text(title)

    if re.search(r"新风|鲜净新风|健康新风|智新风|静新风|增氧新风|大新风量", text):
        return 1

    return 0


def extract_energy_saving(title):
    text = clean_text(title)

    if re.search(
        r"省电|节能|酷省电|巨省电|真省电|超省电|净省电|易省电|智省电|静省电|省电侠|节能王子|AI省电|一键酷省电",
        text,
        re.I
    ):
        return 1

    return 0


wb = load_workbook(excel_file)
ws = wb.active

ws["R1"] = "匹数"
ws["S1"] = "机型"
ws["T1"] = "冷暖型和单冷型"
ws["U1"] = "新风功能"
ws["V1"] = "是否为省电型"

for row in range(2, ws.max_row + 1):
    print(f"正在处理第 {row} 行 / 共 {ws.max_row} 行")

    title = ws[f"L{row}"].value

    ws[f"R{row}"] = extract_hp(title)
    ws[f"S{row}"] = extract_machine_type(title)
    ws[f"T{row}"] = extract_cold_warm_type(title)
    ws[f"U{row}"] = extract_fresh_air(title)
    ws[f"V{row}"] = extract_energy_saving(title)

wb.save(excel_file)

print("处理完成，结果已直接写入原 Excel 文件")
print("文件路径：", excel_file)