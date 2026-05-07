import re
import time
import ctypes
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright


# =========================
# 改成你的 Excel 路径
# 如果第1行是表头，就改成 2
# =========================
excel_file = r"C:\Users\15322\Desktop\f.xlsx"
start_row = 1


# =========================
# 页面注入脚本
# 右键 = 提取
# Ctrl + 右键 = 跳过
# =========================
HOTKEY_JS = r"""
(() => {
  if (window.__jd_hotkey_installed__) return;
  window.__jd_hotkey_installed__ = true;
  window.__jd_extract_signal__ = "";

  document.addEventListener("contextmenu", function (e) {
    e.preventDefault();
    e.stopPropagation();

    if (e.ctrlKey) {
      window.__jd_extract_signal__ = "skip";
    } else {
      window.__jd_extract_signal__ = "extract";
    }
  }, true);
})();
"""


def copy_to_clipboard(text):
    text = str(text)

    CF_UNICODETEXT = 13
    GHND = 0x0042  # GMEM_MOVEABLE | GMEM_ZEROINIT

    user32 = ctypes.windll.user32
    kernel32 = ctypes.windll.kernel32

    user32.OpenClipboard.argtypes = [ctypes.c_void_p]
    user32.OpenClipboard.restype = ctypes.c_int

    user32.EmptyClipboard.argtypes = []
    user32.EmptyClipboard.restype = ctypes.c_int

    user32.SetClipboardData.argtypes = [ctypes.c_uint, ctypes.c_void_p]
    user32.SetClipboardData.restype = ctypes.c_void_p

    user32.CloseClipboard.argtypes = []
    user32.CloseClipboard.restype = ctypes.c_int

    kernel32.GlobalAlloc.argtypes = [ctypes.c_uint, ctypes.c_size_t]
    kernel32.GlobalAlloc.restype = ctypes.c_void_p

    kernel32.GlobalLock.argtypes = [ctypes.c_void_p]
    kernel32.GlobalLock.restype = ctypes.c_void_p

    kernel32.GlobalUnlock.argtypes = [ctypes.c_void_p]
    kernel32.GlobalUnlock.restype = ctypes.c_int

    kernel32.GlobalFree.argtypes = [ctypes.c_void_p]
    kernel32.GlobalFree.restype = ctypes.c_void_p

    data = text.encode("utf-16-le") + b"\x00\x00"
    h_global = None

    for _ in range(10):
        if user32.OpenClipboard(None):
            break
        time.sleep(0.1)
    else:
        raise OSError("无法打开剪贴板，可能被其他程序占用")

    try:
        if not user32.EmptyClipboard():
            raise OSError("EmptyClipboard 失败")

        h_global = kernel32.GlobalAlloc(GHND, len(data))
        if not h_global:
            raise MemoryError("GlobalAlloc 失败")

        p_global = kernel32.GlobalLock(h_global)
        if not p_global:
            kernel32.GlobalFree(h_global)
            raise MemoryError("GlobalLock 失败")

        ctypes.memmove(p_global, data, len(data))
        kernel32.GlobalUnlock(h_global)

        if not user32.SetClipboardData(CF_UNICODETEXT, h_global):
            kernel32.GlobalFree(h_global)
            raise OSError("SetClipboardData 失败")

        h_global = None

    finally:
        user32.CloseClipboard()


def extract_first(text, patterns):
    for pattern in patterns:
        m = re.search(pattern, text, flags=re.I)
        if m:
            return m.group(1)
    return None


def extract_ids_from_html(html, url):
    # =========================
    # shopId 提取规则
    # 优先从当前网址中提取：
    # https://mall.jd.com/index-13616961.html?from=pc
    # 提取结果：13616961
    # =========================
    shop_id = None

    m = re.search(r'index-(\d+)\.html', url, flags=re.I)
    if m:
        shop_id = m.group(1)

    # 如果网址没有提取到，再从隐藏 input 中提取
    # 兼容：
    # <input type="hidden" value="13616961" id="shop_id">
    # <input type="hidden" id="shop_id" value="13616961">
    if not shop_id:
        shop_id = extract_first(
            html,
            [
                r'<input[^>]*id=["\']shop_id["\'][^>]*value=["\'](\d+)["\']',
                r'<input[^>]*value=["\'](\d+)["\'][^>]*id=["\']shop_id["\']'
            ]
        )

    # =========================
    # venderId 提取规则
    # 从页面 HTML 中提取：
    # venderId=14740429
    # 提取结果：14740429
    # =========================
    vender_id = extract_first(
        html,
        [
            r'venderId=(\d+)'
        ]
    )

    return vender_id, shop_id


def inject_script_to_existing_pages(context):
    for page in context.pages:
        try:
            page.evaluate(HOTKEY_JS)
        except Exception:
            pass


def wait_for_signal(context, poll_seconds=0.5):
    while True:
        pages = list(context.pages)

        for page in reversed(pages):
            if page.is_closed():
                continue

            try:
                signal = page.evaluate("""
                    () => {
                        const v = window.__jd_extract_signal__ || "";
                        if (v) {
                            window.__jd_extract_signal__ = "";
                        }
                        return v;
                    }
                """)
            except Exception:
                continue

            if signal in ("extract", "skip"):
                return page, signal

        time.sleep(poll_seconds)


def main():
    wb = load_workbook(excel_file)
    ws = wb.active

    rows_to_do = []

    for row in range(start_row, ws.max_row + 1):
        key = ws.cell(row=row, column=1).value
        b_val = ws.cell(row=row, column=2).value
        c_val = ws.cell(row=row, column=3).value

        if key is None or str(key).strip() == "":
            continue

        # 只处理 B、C 为空的行
        # 如果你想全部重跑，把下面这个判断删掉
        if b_val not in (None, "") and c_val not in (None, ""):
            continue

        rows_to_do.append((row, str(key).strip()))

    print(f"待处理关键词数量：{len(rows_to_do)}")

    with sync_playwright() as p:
        browser = p.chromium.connect_over_cdp("http://127.0.0.1:9222")

        if browser.contexts:
            context = browser.contexts[0]
        else:
            context = browser.new_context()

        # 给后续新开的页面自动注入
        context.add_init_script(HOTKEY_JS)

        # 给当前已打开的页面补注入
        inject_script_to_existing_pages(context)

        for row_idx, key in rows_to_do:
            print("\n" + "=" * 60)
            print(f"当前处理第 {row_idx} 行")
            print(f"A列关键词：{key}")

            try:
                copy_to_clipboard(key)
                print(f"已复制到剪贴板：{key}")
            except Exception as e:
                print(f"复制到剪贴板失败：{e}")

            print("请到 Edge 中 Ctrl+V 搜索并点击进店。")
            print("进入目标店铺页后：右键一下 = 提取")
            print("如果要跳过当前行：Ctrl + 右键")

            page, signal = wait_for_signal(context)
            page.bring_to_front()

            if signal == "skip":
                print(f"第 {row_idx} 行已跳过。")
                ws.cell(row=row_idx, column=2).value = None
                ws.cell(row=row_idx, column=3).value = None
                wb.save(excel_file)
                continue

            try:
                html = page.content()
                url = page.url

                vender_id, shop_id = extract_ids_from_html(html, url)

                print(f"当前页面标题：{page.title()}")
                print(f"当前页面网址：{url}")
                print("提取结果：")
                print("venderId =", vender_id)
                print("shopId   =", shop_id)

                # B列写 venderId
                # C列写 shopId
                ws.cell(row=row_idx, column=2).value = vender_id
                ws.cell(row=row_idx, column=3).value = shop_id
                wb.save(excel_file)

            except Exception as e:
                print(f"第 {row_idx} 行提取失败：{e}")
                ws.cell(row=row_idx, column=2).value = None
                ws.cell(row=row_idx, column=3).value = None
                wb.save(excel_file)

        wb.save(excel_file)
        print("\n全部处理完成。")


if __name__ == "__main__":
    main()