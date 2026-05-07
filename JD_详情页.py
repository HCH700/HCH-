import re
import time
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError


excel_file = r"C:\Users\15322\Desktop\f.xlsx"
output_file = r"C:\Users\15322\Desktop\f_结果.xlsx"

start_row = 1
sleep_seconds = 3


def clean_text(text):
    if text is None:
        return None
    return re.sub(r"\s+", " ", text).strip()


def get_date_text(page):
    try:
        elements = page.locator("taro-view-core.shopx_onvQO1.adaptive-box-flex")
        count = elements.count()

        if count >= 3:
            return elements.nth(2).inner_text().strip()
        else:
            return None
    except Exception:
        return None


def main():
    wb = load_workbook(excel_file)
    ws = wb.active

    with sync_playwright() as p:
        browser = p.chromium.launch(channel="msedge", headless=False)
        context = browser.new_context()
        page = context.new_page()

        for row in range(start_row, ws.max_row + 1):
            shopid = ws.cell(row=row, column=3).value
            venderid = ws.cell(row=row, column=2).value

            if shopid is None or venderid is None:
                continue

            shopid = str(shopid).strip()
            venderid = str(venderid).strip()

            if shopid == "" or venderid == "":
                continue

            url = f"https://shop.m.jd.com/shop/introduce?shopId={shopid}&venderId={venderid}"
            ws.cell(row=row, column=5).value = url

            print(f"\n正在处理第 {row} 行")
            print(url)

            try:
                page.goto(url, wait_until="domcontentloaded", timeout=60000)
                page.wait_for_timeout(2500)

                title = page.title()
                body_text = clean_text(page.locator("body").inner_text())
                date_text = get_date_text(page)

                ws.cell(row=row, column=6).value = title
                ws.cell(row=row, column=7).value = body_text[:2000] if body_text else None
                ws.cell(row=row, column=8).value = "成功"
                ws.cell(row=row, column=9).value = date_text   # I列：日期

                print("页面标题：", title)
                print("日期：", date_text)

                wb.save(output_file)
                print("成功")

            except PlaywrightTimeoutError:
                ws.cell(row=row, column=8).value = "超时"
                ws.cell(row=row, column=9).value = None
                wb.save(output_file)
                print("超时")

            except Exception as e:
                ws.cell(row=row, column=8).value = f"失败: {str(e)[:100]}"
                ws.cell(row=row, column=9).value = None
                wb.save(output_file)
                print("失败：", e)

            time.sleep(sleep_seconds)

        browser.close()

    wb.save(output_file)
    print(f"\n全部完成，结果已保存到：{output_file}")


if __name__ == "__main__":
    main()